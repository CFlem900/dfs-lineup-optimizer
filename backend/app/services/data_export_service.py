"""Analytical data export service for DFS auditing and research.

Exports enriched player pool metrics and game environment summaries
to CSV and XLSX formats for manual analysis.  The XLSX version includes
conditional formatting, frozen headers, and auto-sized columns.

Usage
-----
    from app.services.data_export_service import DataExportService

    exporter = DataExportService()
    exporter.export_player_metrics(player_pool)
    exporter.export_game_environments(games)
    # Or both at once:
    exporter.export_all(player_pool, games)
"""

import logging
import os
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd

logger = logging.getLogger(__name__)

# Default output directory (relative to backend root)
DEFAULT_EXPORT_DIR = "exports"


class DataExportService:
    """Export player projections and game environments to CSV/XLSX."""

    def __init__(self, output_dir: str = DEFAULT_EXPORT_DIR):
        self.output_dir = Path(output_dir)

    def _ensure_dir(self) -> Path:
        """Create the export directory if it doesn't exist."""
        self.output_dir.mkdir(parents=True, exist_ok=True)
        return self.output_dir

    # ──────────────────────────────────────────────────────────────
    # Player Metrics Export
    # ──────────────────────────────────────────────────────────────

    def export_player_metrics(
        self,
        player_pool: List[Any],
        filename: str = "player_metrics_report",
    ) -> Dict[str, str]:
        """Export the enriched player pool to CSV and XLSX.

        Parameters
        ----------
        player_pool : list
            List of PlayerPoolEntry objects (or any object with the
            standard projection fields).
        filename : str
            Base filename (without extension).

        Returns
        -------
        dict
            {"csv": "/path/to/file.csv", "xlsx": "/path/to/file.xlsx"}
        """
        out = self._ensure_dir()

        rows = []
        for p in player_pool:
            proj_fp = float(getattr(p, "projected_fp", 0.0) or 0.0)
            _raw_sal = getattr(p, "salary", 0) or 0
            salary = int(float(_raw_sal)) if _raw_sal else 0
            proj_min = float(getattr(p, "projected_minutes", 0.0) or 0.0)

            # Compute FPPM (base and adjusted)
            base_fppm = round(proj_fp / proj_min, 3) if proj_min > 0 else 0.0

            # Value score: projected_fp / (salary / 1000)
            value = round(proj_fp / (salary / 1000), 2) if salary > 0 else 0.0

            # Leverage = optimal% - ownership%
            optimal_pct = getattr(p, "sim_optimal_pct", None)
            if optimal_pct is None:
                # Fall back to boom_probability as a proxy
                optimal_pct = getattr(p, "boom_probability", None)
            ownership_pct = getattr(p, "estimated_ownership", None)

            leverage = None
            if optimal_pct is not None and ownership_pct is not None:
                leverage = round(optimal_pct - ownership_pct, 2)

            rows.append({
                "Name": getattr(p, "player_name", ""),
                "Position": getattr(p, "position", ""),
                "Team": getattr(p, "team_abbreviation", ""),
                "Opponent": getattr(p, "opponent_abbreviation", "") or "",
                "Salary": salary,
                "Proj_Minutes": round(proj_min, 1),
                "Base_FPPM": round(base_fppm, 3),
                "Proj_FP": round(proj_fp, 2),
                "Floor_FP": round(getattr(p, "floor_fp", 0.0) or 0.0, 2),
                "Ceiling_FP": round(getattr(p, "ceiling_fp", 0.0) or 0.0, 2),
                "Value": value,
                "Ownership_Pct": round(ownership_pct, 1) if ownership_pct is not None else None,
                "Optimal_Pct": round(optimal_pct, 1) if optimal_pct is not None else None,
                "Leverage": leverage,
                "Sim_Floor_P10": round(getattr(p, "sim_p10", 0.0) or 0.0, 2),
                "Sim_Median_P50": round(getattr(p, "sim_p50", 0.0) or 0.0, 2),
                "Sim_Ceiling_P90": round(getattr(p, "sim_p90", 0.0) or 0.0, 2),
                "Sim_StdDev": round(getattr(p, "sim_std", 0.0) or 0.0, 2),
                "Vegas_Spread": getattr(p, "vegas_spread", None),
                "Game_Total": getattr(p, "game_total", None),
                "Injury": getattr(p, "injury_status", None) or "",
                "Noise_Archetype": getattr(p, "noise_archetype", None) or "",
                "Proj_Source": getattr(p, "projection_source", None) or "rotation",
            })

        df = pd.DataFrame(rows)

        # Sort by Leverage descending (best tournament plays first),
        # with nulls at the bottom.  Coerce to numeric first to handle
        # mixed types from upstream (str salaries, None values, etc.).
        df["Leverage"] = pd.to_numeric(df["Leverage"], errors="coerce")
        df = df.sort_values(
            by="Leverage", ascending=False, na_position="last"
        ).reset_index(drop=True)

        # Export CSV
        csv_path = out / f"{filename}.csv"
        df.to_csv(csv_path, index=False)

        # Export XLSX with formatting
        xlsx_path = out / f"{filename}.xlsx"
        self._write_formatted_xlsx(df, xlsx_path, sheet_name="Player Metrics")

        logger.info(
            "[DataExport] Player metrics exported: %d players → %s (.csv + .xlsx)",
            len(df), out,
        )
        return {"csv": str(csv_path), "xlsx": str(xlsx_path)}

    # ──────────────────────────────────────────────────────────────
    # Game Environments Export
    # ──────────────────────────────────────────────────────────────

    def export_game_environments(
        self,
        games: List[Any],
        filename: str = "game_environments_report",
    ) -> Dict[str, str]:
        """Export game environment summaries to CSV and XLSX.

        Parameters
        ----------
        games : list
            List of GameInfo objects (or dicts with game fields).

        Returns
        -------
        dict
            {"csv": "/path/to/file.csv", "xlsx": "/path/to/file.xlsx"}
        """
        out = self._ensure_dir()

        rows = []
        for g in games:
            # Support both GameInfo objects and dicts
            _get = (
                (lambda k, d=None: g.get(k, d))
                if isinstance(g, dict)
                else (lambda k, d=None: getattr(g, k, d))
            )

            home = _get("home_team")
            away = _get("away_team")
            home_abbr = (
                getattr(home, "team_abbreviation", "???") if home else "???"
            )
            away_abbr = (
                getattr(away, "team_abbreviation", "???") if away else "???"
            )

            vegas_total = _get("over_under") or _get("projected_total", 0.0)
            vegas_spread = _get("vegas_spread") or _get("projected_spread", 0.0)
            pace = _get("projected_pace", 0.0)
            pace_label = _get("pace_label", "")
            tip_time = _get("game_time_et", "")

            # Blowout risk: |spread| >= 13
            abs_spread = abs(vegas_spread) if vegas_spread else 0
            blowout_risk = abs_spread >= 13.0

            rows.append({
                "Game": f"{away_abbr} @ {home_abbr}",
                "Tip_Time_ET": tip_time or "",
                "Vegas_Total": round(vegas_total, 1) if vegas_total else None,
                "Vegas_Spread": round(vegas_spread, 1) if vegas_spread else None,
                "Proj_Total": round(_get("projected_total", 0.0) or 0.0, 1),
                "Proj_Pace": round(pace, 1) if pace else None,
                "Pace_Label": pace_label,
                "Home_Implied": round(
                    _get("projected_home_score", 0.0) or 0.0, 1
                ),
                "Away_Implied": round(
                    _get("projected_away_score", 0.0) or 0.0, 1
                ),
                "Blowout_Risk": blowout_risk,
                "Spread_Abs": round(abs_spread, 1),
            })

        df = pd.DataFrame(rows)

        # Sort by Vegas Total descending (highest-scoring games first)
        if "Vegas_Total" in df.columns and not df.empty:
            df = df.sort_values(
                by="Vegas_Total", ascending=False, na_position="last"
            ).reset_index(drop=True)

        csv_path = out / f"{filename}.csv"
        df.to_csv(csv_path, index=False)

        xlsx_path = out / f"{filename}.xlsx"
        self._write_formatted_xlsx(df, xlsx_path, sheet_name="Game Environments")

        logger.info(
            "[DataExport] Game environments exported: %d games → %s (.csv + .xlsx)",
            len(df), out,
        )
        return {"csv": str(csv_path), "xlsx": str(xlsx_path)}

    # ──────────────────────────────────────────────────────────────
    # Combo Export
    # ──────────────────────────────────────────────────────────────

    def export_all(
        self,
        player_pool: List[Any],
        games: Optional[List[Any]] = None,
        timestamp: bool = True,
    ) -> Dict[str, Dict[str, str]]:
        """Export both player metrics and game environments.

        When *timestamp* is True, appends a UTC timestamp to filenames
        so each run produces unique files for historical comparison.
        """
        suffix = ""
        if timestamp:
            suffix = f"_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}"

        result = {}
        result["players"] = self.export_player_metrics(
            player_pool, filename=f"player_metrics{suffix}"
        )
        if games:
            result["games"] = self.export_game_environments(
                games, filename=f"game_environments{suffix}"
            )

        logger.info(
            "[DataExport] Full export complete → %s", self.output_dir,
        )
        return result

    # ──────────────────────────────────────────────────────────────
    # XLSX Formatting
    # ──────────────────────────────────────────────────────────────

    @staticmethod
    def _write_formatted_xlsx(
        df: pd.DataFrame,
        path: Path,
        sheet_name: str = "Sheet1",
    ) -> None:
        """Write a DataFrame to XLSX with professional formatting.

        Features:
          - Frozen header row
          - Auto-sized columns (capped at 30 chars)
          - Bold header with dark fill
          - Conditional formatting: green for positive Leverage,
            red for negative; red for Blowout_Risk = True
        """
        try:
            from openpyxl.styles import Alignment, Font, PatternFill, numbers
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.warning(
                "[DataExport] openpyxl not installed — falling back to plain XLSX"
            )
            df.to_excel(path, index=False, sheet_name=sheet_name)
            return

        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            ws = writer.sheets[sheet_name]

            # ── Header formatting ──────────────────────────────────
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(
                start_color="2F5496", end_color="2F5496", fill_type="solid"
            )
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            # ── Freeze top row ─────────────────────────────────────
            ws.freeze_panes = "A2"

            # ── Auto-size columns ──────────────────────────────────
            for col_idx in range(1, ws.max_column + 1):
                col_letter = get_column_letter(col_idx)
                header_len = len(str(ws.cell(row=1, column=col_idx).value or ""))
                max_len = header_len
                # Sample first 50 data rows for width
                for row_idx in range(2, min(ws.max_row + 1, 52)):
                    val = ws.cell(row=row_idx, column=col_idx).value
                    if val is not None:
                        max_len = max(max_len, len(str(val)))
                ws.column_dimensions[col_letter].width = min(max_len + 3, 30)

            # ── Conditional formatting for Leverage column ─────────
            col_names = [
                ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)
            ]
            if "Leverage" in col_names:
                lev_col = col_names.index("Leverage") + 1
                green_font = Font(color="006100")
                green_fill = PatternFill(
                    start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
                )
                red_font = Font(color="9C0006")
                red_fill = PatternFill(
                    start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                )
                for row_idx in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row_idx, column=lev_col)
                    if cell.value is not None:
                        try:
                            _val = float(cell.value)
                        except (ValueError, TypeError):
                            continue
                        if _val > 0:
                            cell.font = green_font
                            cell.fill = green_fill
                        elif _val < -5:
                            cell.font = red_font
                            cell.fill = red_fill

            # ── Highlight blowout risk rows ────────────────────────
            if "Blowout_Risk" in col_names:
                risk_col = col_names.index("Blowout_Risk") + 1
                orange_fill = PatternFill(
                    start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
                )
                for row_idx in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row_idx, column=risk_col)
                    if cell.value is True or str(cell.value).lower() == "true":
                        for c in range(1, ws.max_column + 1):
                            ws.cell(row=row_idx, column=c).fill = orange_fill

            # ── Dollar format for Salary column ────────────────────
            if "Salary" in col_names:
                sal_col = col_names.index("Salary") + 1
                for row_idx in range(2, ws.max_row + 1):
                    ws.cell(row=row_idx, column=sal_col).number_format = (
                        numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                    )
